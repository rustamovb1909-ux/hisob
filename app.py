# -*- coding: utf-8 -*-
"""
Hisobchi — Flask backend
- Telegram WebApp initData tekshiruvi (xavfsizlik)
- Faqat botda ro'yxatdan o'tgan (raqam yuborgan) userlar API dan foydalana oladi
- PostgreSQL (Render Postgres) bilan ishlaydi
"""
import os
import hmac
import hashlib
import json
import io
from datetime import datetime
from urllib.parse import parse_qsl

import psycopg2
import psycopg2.extras
import requests
from flask import Flask, request, jsonify, render_template, send_file

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER

BOT_TOKEN = os.environ.get("BOT_TOKEN", "")
BOT_USERNAME = os.environ.get("BOT_USERNAME", "")  # masalan: mening_hisobchi_bot (@ belgisisiz)
DATABASE_URL = os.environ.get("DATABASE_URL", "")

app = Flask(__name__)


# ---------------------------------------------------------------------------
# Ma'lumotlar bazasi
# ---------------------------------------------------------------------------
def get_conn():
    """Har chaqiriqda yangi ulanish (Render'da connection pool shart emas,
    lekin xohlasangiz keyinchalik psycopg2.pool ga o'tkazish oson)."""
    conn = psycopg2.connect(DATABASE_URL, sslmode="require")
    return conn


def init_db():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            telegram_id BIGINT PRIMARY KEY,
            phone TEXT NOT NULL,
            first_name TEXT,
            last_name TEXT,
            username TEXT,
            created_at TIMESTAMP DEFAULT NOW()
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS transactions (
            id SERIAL PRIMARY KEY,
            telegram_id BIGINT NOT NULL REFERENCES users(telegram_id) ON DELETE CASCADE,
            type TEXT NOT NULL CHECK (type IN ('income', 'expense')),
            amount NUMERIC NOT NULL,
            category TEXT DEFAULT '',
            note TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT NOW()
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_tx_user ON transactions(telegram_id)")
    cur.execute("""
        CREATE TABLE IF NOT EXISTS debts (
            id SERIAL PRIMARY KEY,
            telegram_id BIGINT NOT NULL REFERENCES users(telegram_id) ON DELETE CASCADE,
            direction TEXT NOT NULL CHECK (direction IN ('given', 'taken')),
            person_name TEXT NOT NULL,
            amount NUMERIC NOT NULL,
            note TEXT DEFAULT '',
            is_paid BOOLEAN DEFAULT FALSE,
            is_payment BOOLEAN DEFAULT FALSE,
            created_at TIMESTAMP DEFAULT NOW(),
            paid_at TIMESTAMP
        )
    """)
    # Eski (avvalroq yaratilgan) jadvallarda is_payment ustuni bo'lmasligi
    # mumkin — CREATE TABLE IF NOT EXISTS uni qo'shmaydi, shuning uchun
    # mavjud bo'lmasa qo'shib qo'yamiz (xavfsiz, ma'lumotlarni o'chirmaydi).
    cur.execute("ALTER TABLE debts ADD COLUMN IF NOT EXISTS is_payment BOOLEAN DEFAULT FALSE")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_debts_user ON debts(telegram_id)")
    conn.commit()
    cur.close()
    conn.close()


def get_user(telegram_id):
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM users WHERE telegram_id = %s", (telegram_id,))
    row = cur.fetchone()
    cur.close()
    conn.close()
    return row


def upsert_user(telegram_id, phone, first_name, last_name, username):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO users (telegram_id, phone, first_name, last_name, username)
        VALUES (%s, %s, %s, %s, %s)
        ON CONFLICT (telegram_id) DO UPDATE
        SET phone = EXCLUDED.phone,
            first_name = EXCLUDED.first_name,
            last_name = EXCLUDED.last_name,
            username = EXCLUDED.username
    """, (telegram_id, phone, first_name, last_name, username))
    conn.commit()
    cur.close()
    conn.close()


def get_transactions(telegram_id):
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        "SELECT * FROM transactions WHERE telegram_id = %s ORDER BY created_at DESC",
        (telegram_id,)
    )
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows


def add_transaction(telegram_id, type_, amount, category, note):
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        INSERT INTO transactions (telegram_id, type, amount, category, note)
        VALUES (%s, %s, %s, %s, %s)
        RETURNING *
    """, (telegram_id, type_, amount, category, note))
    row = cur.fetchone()
    conn.commit()
    cur.close()
    conn.close()
    return row


def get_transaction(telegram_id, tx_id):
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        "SELECT * FROM transactions WHERE id = %s AND telegram_id = %s",
        (tx_id, telegram_id)
    )
    row = cur.fetchone()
    cur.close()
    conn.close()
    return row


def delete_transaction(telegram_id, tx_id):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "DELETE FROM transactions WHERE id = %s AND telegram_id = %s",
        (tx_id, telegram_id)
    )
    deleted = cur.rowcount
    conn.commit()
    cur.close()
    conn.close()
    return deleted > 0


def update_transaction(telegram_id, tx_id, type_, amount, category, note):
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        UPDATE transactions
        SET type = %s, amount = %s, category = %s, note = %s
        WHERE id = %s AND telegram_id = %s
        RETURNING *
    """, (type_, amount, category, note, tx_id, telegram_id))
    row = cur.fetchone()
    conn.commit()
    cur.close()
    conn.close()
    return row


def get_monthly_summary(telegram_id):
    """Joriy oy (UTC bo'yicha) uchun kirim/xarajat yig'indisi va xarajat
    kategoriyalari bo'yicha taqsimotni qaytaradi. Bot orqali "Oylik hisobot"
    tugmasi bosilganda ishlatiladi."""
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT type, category, SUM(amount) AS total
        FROM transactions
        WHERE telegram_id = %s
          AND date_trunc('month', created_at) = date_trunc('month', NOW())
        GROUP BY type, category
        ORDER BY total DESC
    """, (telegram_id,))
    rows = cur.fetchall()
    cur.close()
    conn.close()

    income = 0.0
    expense = 0.0
    categories = []
    for r in rows:
        total = float(r["total"])
        if r["type"] == "income":
            income += total
        else:
            expense += total
            categories.append((r["category"], total))

    return {
        "income": income,
        "expense": expense,
        "balance": income - expense,
        "categories": categories,
    }


# ---------------------------------------------------------------------------
# Qarz daftari
# ---------------------------------------------------------------------------
def get_debts(telegram_id):
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        "SELECT * FROM debts WHERE telegram_id = %s ORDER BY is_paid ASC, created_at DESC",
        (telegram_id,)
    )
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows


def add_debt(telegram_id, direction, person_name, amount, note, is_payment=False):
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        INSERT INTO debts (telegram_id, direction, person_name, amount, note, is_payment)
        VALUES (%s, %s, %s, %s, %s, %s)
        RETURNING *
    """, (telegram_id, direction, person_name, amount, note, is_payment))
    row = cur.fetchone()
    conn.commit()
    cur.close()
    conn.close()
    return row


def get_debt(telegram_id, debt_id):
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        "SELECT * FROM debts WHERE id = %s AND telegram_id = %s",
        (debt_id, telegram_id)
    )
    row = cur.fetchone()
    cur.close()
    conn.close()
    return row


def update_debt(telegram_id, debt_id, person_name, amount, note):
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        UPDATE debts
        SET person_name = %s, amount = %s, note = %s
        WHERE id = %s AND telegram_id = %s
        RETURNING *
    """, (person_name, amount, note, debt_id, telegram_id))
    row = cur.fetchone()
    conn.commit()
    cur.close()
    conn.close()
    return row


def set_debt_paid(telegram_id, debt_id, is_paid):
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        UPDATE debts
        SET is_paid = %s, paid_at = CASE WHEN %s THEN NOW() ELSE NULL END
        WHERE id = %s AND telegram_id = %s
        RETURNING *
    """, (is_paid, is_paid, debt_id, telegram_id))
    row = cur.fetchone()
    conn.commit()
    cur.close()
    conn.close()
    return row


def delete_debt(telegram_id, debt_id):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "DELETE FROM debts WHERE id = %s AND telegram_id = %s",
        (debt_id, telegram_id)
    )
    deleted = cur.rowcount
    conn.commit()
    cur.close()
    conn.close()
    return deleted > 0


# ---------------------------------------------------------------------------
# Excel / PDF hisobot generatsiyasi
# ---------------------------------------------------------------------------
UZ_MONTHS = [
    "yanvar", "fevral", "mart", "aprel", "may", "iyun",
    "iyul", "avgust", "sentyabr", "oktyabr", "noyabr", "dekabr",
]

BRAND_DARK = "0A0E27"
BRAND_GREEN = "00B894"
BRAND_RED = "FF4757"


def _monthly_breakdown(transactions):
    """Tranzaksiyalarni oy bo'yicha guruhlab, har oy uchun kirim/xarajat/
    balansni hisoblaydi (eng yangi oy birinchi)."""
    months = {}
    for t in transactions:
        d = t["created_at"]
        key = (d.year, d.month)
        if key not in months:
            months[key] = {"year": d.year, "month": d.month, "income": 0.0, "expense": 0.0}
        if t["type"] == "income":
            months[key]["income"] += float(t["amount"])
        else:
            months[key]["expense"] += float(t["amount"])
    return sorted(months.values(), key=lambda m: (m["year"], m["month"]), reverse=True)


def generate_excel_report(telegram_id, display_name):
    transactions = get_transactions(telegram_id)
    debts = get_debts(telegram_id)

    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color=BRAND_DARK, end_color=BRAND_DARK, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=14, color=BRAND_DARK)
    thin_border = Border(*(Side(style="thin", color="DDDDDD"),) * 4)

    def style_header_row(ws, row_num, ncols):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def autosize(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # --- 1-bet: Oylik xulosa ---
    ws1 = wb.active
    ws1.title = "Oylik xulosa"
    ws1["A1"] = "Hisobchi — Moliyaviy hisobot"
    ws1["A1"].font = title_font
    ws1["A2"] = f"Foydalanuvchi: {display_name}"
    ws1["A3"] = f"Yaratildi: {datetime.utcnow().strftime('%d.%m.%Y %H:%M')} (UTC)"

    headers = ["Oy", "Kirim (so'm)", "Xarajat (so'm)", "Balans (so'm)"]
    ws1.append([])
    ws1.append(headers)
    style_header_row(ws1, 5, len(headers))

    for m in _monthly_breakdown(transactions):
        balance = m["income"] - m["expense"]
        label = f"{UZ_MONTHS[m['month'] - 1]} {m['year']}"
        ws1.append([label, m["income"], m["expense"], balance])

    for row in ws1.iter_rows(min_row=6, max_row=ws1.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border
            if cell.column > 1:
                cell.number_format = "#,##0"

    autosize(ws1, [22, 18, 18, 18])

    # --- 2-bet: Barcha tranzaksiyalar ---
    ws2 = wb.create_sheet("Tranzaksiyalar")
    headers2 = ["Sana", "Turi", "Summa (so'm)", "Kategoriya", "Izoh"]
    ws2.append(headers2)
    style_header_row(ws2, 1, len(headers2))

    for t in transactions:
        ws2.append([
            t["created_at"].strftime("%d.%m.%Y %H:%M"),
            "Kirim" if t["type"] == "income" else "Xarajat",
            float(t["amount"]),
            t["category"] or "",
            t["note"] or "",
        ])

    for row in ws2.iter_rows(min_row=2, max_row=max(ws2.max_row, 2), min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border
            if cell.column == 3:
                cell.number_format = "#,##0"

    autosize(ws2, [18, 12, 16, 20, 30])

    # --- 3-bet: Qarz daftari ---
    ws3 = wb.create_sheet("Qarz daftari")
    headers3 = ["Ism", "Yo'nalish", "Summa (so'm)", "Holat", "Sana", "Izoh"]
    ws3.append(headers3)
    style_header_row(ws3, 1, len(headers3))

    for d in debts:
        ws3.append([
            d["person_name"],
            "Menga qarzdor" if d["direction"] == "given" else "Men qarzdorman",
            float(d["amount"]),
            "To'landi" if d["is_paid"] else "To'lanmagan",
            d["created_at"].strftime("%d.%m.%Y"),
            d["note"] or "",
        ])

    for row in ws3.iter_rows(min_row=2, max_row=max(ws3.max_row, 2), min_col=1, max_col=6):
        for cell in row:
            cell.border = thin_border
            if cell.column == 3:
                cell.number_format = "#,##0"

    autosize(ws3, [18, 16, 16, 14, 14, 30])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_pdf_report(telegram_id, display_name):
    transactions = get_transactions(telegram_id)
    debts = get_debts(telegram_id)
    months = _monthly_breakdown(transactions)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=18 * mm, bottomMargin=18 * mm,
        leftMargin=16 * mm, rightMargin=16 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleUz", parent=styles["Title"], fontSize=18,
        textColor=colors.HexColor("#0A0E27"), alignment=TA_CENTER, spaceAfter=4,
    )
    meta_style = ParagraphStyle(
        "MetaUz", parent=styles["Normal"], fontSize=9,
        textColor=colors.HexColor("#666666"), alignment=TA_CENTER, spaceAfter=14,
    )
    section_style = ParagraphStyle(
        "SectionUz", parent=styles["Heading2"], fontSize=13,
        textColor=colors.HexColor("#0A0E27"), spaceBefore=14, spaceAfter=8,
    )

    def fmt_num(n):
        return f"{n:,.0f}".replace(",", " ")

    elements = [
        Paragraph("Hisobchi — Moliyaviy hisobot", title_style),
        Paragraph(
            f"Foydalanuvchi: {display_name} &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"Yaratildi: {datetime.utcnow().strftime('%d.%m.%Y %H:%M')} (UTC)",
            meta_style,
        ),
    ]

    total_income = sum(float(t["amount"]) for t in transactions if t["type"] == "income")
    total_expense = sum(float(t["amount"]) for t in transactions if t["type"] == "expense")

    elements.append(Paragraph("Umumiy holat", section_style))
    summary_data = [
        ["Jami kirim", "Jami xarajat", "Balans"],
        [f"{fmt_num(total_income)} so'm", f"{fmt_num(total_expense)} so'm",
         f"{fmt_num(total_income - total_expense)} so'm"],
    ]
    summary_table = Table(summary_data, colWidths=[55 * mm, 55 * mm, 55 * mm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A0E27")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
    ]))
    elements.append(summary_table)

    if months:
        elements.append(Paragraph("Oylik xulosa", section_style))
        month_rows = [["Oy", "Kirim", "Xarajat", "Balans"]]
        for m in months:
            balance = m["income"] - m["expense"]
            month_rows.append([
                f"{UZ_MONTHS[m['month'] - 1]} {m['year']}",
                fmt_num(m["income"]), fmt_num(m["expense"]), fmt_num(balance),
            ])
        month_table = Table(month_rows, colWidths=[45 * mm, 40 * mm, 40 * mm, 40 * mm])
        month_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A0E27")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F6FA")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(month_table)

    if debts:
        elements.append(Paragraph("Qarz daftari", section_style))
        debt_rows = [["Ism", "Yo'nalish", "Summa", "Holat"]]
        for d in debts:
            debt_rows.append([
                d["person_name"],
                "Menga qarzdor" if d["direction"] == "given" else "Men qarzdorman",
                fmt_num(float(d["amount"])),
                "To'landi" if d["is_paid"] else "To'lanmagan",
            ])
        debt_table = Table(debt_rows, colWidths=[45 * mm, 40 * mm, 35 * mm, 45 * mm])
        debt_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A0E27")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (2, 0), (2, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F6FA")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(debt_table)

    if transactions:
        elements.append(Spacer(1, 6))
        elements.append(Paragraph("Barcha tranzaksiyalar", section_style))
        tx_rows = [["Sana", "Turi", "Summa", "Kategoriya"]]
        for t in transactions:
            tx_rows.append([
                t["created_at"].strftime("%d.%m.%Y %H:%M"),
                "Kirim" if t["type"] == "income" else "Xarajat",
                fmt_num(float(t["amount"])),
                t["category"] or "",
            ])
        tx_table = Table(tx_rows, colWidths=[38 * mm, 25 * mm, 32 * mm, 65 * mm], repeatRows=1)
        tx_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A0E27")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (2, 0), (2, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F6FA")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(tx_table)

    doc.build(elements)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Telegram WebApp initData tekshiruvi
# https://core.telegram.org/bots/webapps#validating-data-received-via-the-web-app
# ---------------------------------------------------------------------------
def verify_init_data(init_data: str):
    """initData imzosini tekshiradi. To'g'ri bo'lsa (dict user, ...) qaytaradi,
    noto'g'ri yoki bo'sh bo'lsa None qaytaradi."""
    if not init_data or not BOT_TOKEN:
        return None
    try:
        pairs = dict(parse_qsl(init_data, strict_parsing=True))
    except ValueError:
        return None

    received_hash = pairs.pop("hash", None)
    if not received_hash:
        return None

    data_check_string = "\n".join(
        f"{k}={v}" for k, v in sorted(pairs.items())
    )
    secret_key = hmac.new(
        b"WebAppData", BOT_TOKEN.encode(), hashlib.sha256
    ).digest()
    computed_hash = hmac.new(
        secret_key, data_check_string.encode(), hashlib.sha256
    ).hexdigest()

    if not hmac.compare_digest(computed_hash, received_hash):
        return None

    user_raw = pairs.get("user")
    if not user_raw:
        return None
    try:
        user = json.loads(user_raw)
    except json.JSONDecodeError:
        return None

    return user


def require_auth():
    """Har bir API so'rovi uchun: initData -> telegram user.
    Qaytaradi: (user_dict, None) yoki (None, (response, status))"""
    init_data = request.headers.get("X-Telegram-Init-Data", "")
    user = verify_init_data(init_data)
    if user is None:
        return None, (jsonify({"error": "Telegram orqali kiring"}), 401)
    return user, None


def require_registered():
    """initData to'g'ri VA foydalanuvchi botda ro'yxatdan o'tgan bo'lishi kerak."""
    user, err = require_auth()
    if err:
        return None, err
    db_user = get_user(user["id"])
    if not db_user:
        return None, (jsonify({
            "error": "not_registered",
            "message": "Avval botda ro'yxatdan o'ting"
        }), 403)
    return user, None


# ---------------------------------------------------------------------------
# Sahifalar
# ---------------------------------------------------------------------------
def to_utc_iso(dt):
    """PostgreSQL NOW() qiymati UTC bo'yicha saqlanadi, lekin "timezone
    yo'q" (naive) holatda qaytadi. Bu funksiya ISO satr oxiriga aniq "Z"
    (UTC) belgisini qo'shadi — shunda brauzer sanani noto'g'ri (o'zining
    mahalliy vaqti deb) talqin qilib, kunni siljitib yubormaydi."""
    return dt.isoformat() + "Z"


@app.route("/")
def index():
    return render_template("index.html", bot_username=BOT_USERNAME)


# ---------------------------------------------------------------------------
# API
# ---------------------------------------------------------------------------
@app.route("/api/me")
def api_me():
    user, err = require_auth()
    if err:
        return err
    db_user = get_user(user["id"])
    if not db_user:
        return jsonify({"registered": False}), 200
    return jsonify({
        "registered": True,
        "first_name": db_user["first_name"] or user.get("first_name", ""),
        "phone": db_user["phone"],
    }), 200


@app.route("/api/transactions", methods=["GET"])
def api_list_transactions():
    user, err = require_registered()
    if err:
        return err
    rows = get_transactions(user["id"])
    result = [{
        "id": r["id"],
        "type": r["type"],
        "amount": float(r["amount"]),
        "category": r["category"],
        "note": r["note"],
        "created_at": to_utc_iso(r["created_at"]),
    } for r in rows]
    return jsonify(result)


@app.route("/api/transactions", methods=["POST"])
def api_add_transaction():
    user, err = require_registered()
    if err:
        return err

    data = request.get_json(silent=True) or {}
    type_ = data.get("type")
    amount = data.get("amount")
    category = (data.get("category") or "").strip()[:100]
    note = (data.get("note") or "").strip()[:200]

    if type_ not in ("income", "expense"):
        return jsonify({"error": "noto'g'ri turi"}), 400
    try:
        amount = float(amount)
    except (TypeError, ValueError):
        return jsonify({"error": "noto'g'ri summa"}), 400
    if amount <= 0:
        return jsonify({"error": "summa 0 dan katta bo'lishi kerak"}), 400
    if type_ == "expense" and not category:
        return jsonify({"error": "kategoriya kerak"}), 400

    row = add_transaction(user["id"], type_, amount, category or "kirim", note)
    return jsonify({
        "id": row["id"],
        "type": row["type"],
        "amount": float(row["amount"]),
        "category": row["category"],
        "note": row["note"],
        "created_at": to_utc_iso(row["created_at"]),
    }), 201


@app.route("/api/transactions/<int:tx_id>", methods=["PUT"])
def api_update_transaction(tx_id):
    user, err = require_registered()
    if err:
        return err

    data = request.get_json(silent=True) or {}
    type_ = data.get("type")
    amount = data.get("amount")
    category = (data.get("category") or "").strip()[:100]
    note = (data.get("note") or "").strip()[:200]

    if type_ not in ("income", "expense"):
        return jsonify({"error": "noto'g'ri turi"}), 400
    try:
        amount = float(amount)
    except (TypeError, ValueError):
        return jsonify({"error": "noto'g'ri summa"}), 400
    if amount <= 0:
        return jsonify({"error": "summa 0 dan katta bo'lishi kerak"}), 400
    if type_ == "expense" and not category:
        return jsonify({"error": "kategoriya kerak"}), 400

    row = update_transaction(user["id"], tx_id, type_, amount, category or "kirim", note)
    if not row:
        return jsonify({"error": "topilmadi"}), 404

    return jsonify({
        "id": row["id"],
        "type": row["type"],
        "amount": float(row["amount"]),
        "category": row["category"],
        "note": row["note"],
        "created_at": to_utc_iso(row["created_at"]),
    }), 200


@app.route("/api/transactions/<int:tx_id>", methods=["DELETE"])
def api_delete_transaction(tx_id):
    user, err = require_registered()
    if err:
        return err
    ok = delete_transaction(user["id"], tx_id)
    if not ok:
        return jsonify({"error": "topilmadi"}), 404
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# API — Qarz daftari
# ---------------------------------------------------------------------------
def serialize_debt(row):
    return {
        "id": row["id"],
        "direction": row["direction"],
        "person_name": row["person_name"],
        "amount": float(row["amount"]),
        "note": row["note"],
        "is_paid": row["is_paid"],
        "is_payment": bool(row["is_payment"]),
        "created_at": to_utc_iso(row["created_at"]),
        "paid_at": to_utc_iso(row["paid_at"]) if row["paid_at"] else None,
    }


@app.route("/api/debts", methods=["GET"])
def api_list_debts():
    user, err = require_registered()
    if err:
        return err
    rows = get_debts(user["id"])
    return jsonify([serialize_debt(r) for r in rows])


@app.route("/api/debts", methods=["POST"])
def api_add_debt():
    user, err = require_registered()
    if err:
        return err

    data = request.get_json(silent=True) or {}
    direction = data.get("direction")
    person_name = (data.get("person_name") or "").strip()[:100]
    amount = data.get("amount")
    note = (data.get("note") or "").strip()[:200]
    is_payment = bool(data.get("is_payment", False))

    if direction not in ("given", "taken"):
        return jsonify({"error": "noto'g'ri turi"}), 400
    if not person_name:
        return jsonify({"error": "ism kiritilmagan"}), 400
    try:
        amount = float(amount)
    except (TypeError, ValueError):
        return jsonify({"error": "noto'g'ri summa"}), 400
    if amount <= 0:
        return jsonify({"error": "summa 0 dan katta bo'lishi kerak"}), 400

    row = add_debt(user["id"], direction, person_name, amount, note, is_payment)
    return jsonify(serialize_debt(row)), 201


@app.route("/api/debts/<int:debt_id>", methods=["PUT"])
def api_update_debt(debt_id):
    user, err = require_registered()
    if err:
        return err

    existing = get_debt(user["id"], debt_id)
    if not existing:
        return jsonify({"error": "topilmadi"}), 404

    data = request.get_json(silent=True) or {}

    # Faqat "to'landi" holatini almashtirish so'ralgan bo'lishi mumkin
    if "is_paid" in data and len(data) == 1:
        row = set_debt_paid(user["id"], debt_id, bool(data["is_paid"]))
        return jsonify(serialize_debt(row))

    person_name = (data.get("person_name") or "").strip()[:100]
    amount = data.get("amount")
    note = (data.get("note") or "").strip()[:200]

    if not person_name:
        return jsonify({"error": "ism kiritilmagan"}), 400
    try:
        amount = float(amount)
    except (TypeError, ValueError):
        return jsonify({"error": "noto'g'ri summa"}), 400
    if amount <= 0:
        return jsonify({"error": "summa 0 dan katta bo'lishi kerak"}), 400

    row = update_debt(user["id"], debt_id, person_name, amount, note)
    if not row:
        return jsonify({"error": "topilmadi"}), 404
    return jsonify(serialize_debt(row))


@app.route("/api/debts/<int:debt_id>", methods=["DELETE"])
def api_delete_debt(debt_id):
    user, err = require_registered()
    if err:
        return err
    ok = delete_debt(user["id"], debt_id)
    if not ok:
        return jsonify({"error": "topilmadi"}), 404
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# API — Excel / PDF hisobot
# ---------------------------------------------------------------------------
BOT_TOKEN_ENV = os.environ.get("BOT_TOKEN", "")


def send_telegram_document(chat_id, filename, file_bytes, caption=""):
    """Faylni Telegram Bot API orqali to'g'ridan-to'g'ri foydalanuvchi
    chatiga yuboradi. Telegram ichidagi brauzer (WebView) orqali fayl
    yuklab olish ko'p qurilmalarda ishlamay qolishi mumkin — shuning
    uchun fayl brauzerga emas, bevosita bot chatiga jo'natiladi."""
    url = f"https://api.telegram.org/bot{BOT_TOKEN_ENV}/sendDocument"
    files = {"document": (filename, file_bytes)}
    data = {"chat_id": chat_id, "caption": caption}
    resp = requests.post(url, data=data, files=files, timeout=30)
    resp.raise_for_status()
    result = resp.json()
    if not result.get("ok"):
        raise RuntimeError(result.get("description", "Telegram xatosi"))
    return result


@app.route("/api/export/send", methods=["POST"])
def api_export_send():
    """Hisobotni (Excel yoki PDF) yaratib, to'g'ridan-to'g'ri foydalanuvchi
    bilan botning chatiga jo'natadi. Frontend endi faylni brauzerda
    yuklab olishga urinmaydi (bu Telegram WebView'da ishonchsiz) —
    o'rniga shu endpointni chaqirib, natijani bot chatidan kutadi."""
    user, err = require_registered()
    if err:
        return err

    data = request.get_json(silent=True) or {}
    fmt = data.get("format")
    if fmt not in ("excel", "pdf"):
        return jsonify({"error": "noto'g'ri format"}), 400

    db_user = get_user(user["id"])
    display_name = db_user["first_name"] or user.get("first_name", "Foydalanuvchi")

    try:
        if fmt == "excel":
            buf = generate_excel_report(user["id"], display_name)
            filename = f"hisobchi_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
        else:
            buf = generate_pdf_report(user["id"], display_name)
            filename = f"hisobchi_{datetime.utcnow().strftime('%Y%m%d')}.pdf"

        send_telegram_document(
            user["id"], filename, buf.read(),
            caption="📊 Moliyaviy hisobotingiz tayyor."
        )
    except Exception as e:
        return jsonify({"error": f"hisobotni yuborib bo'lmadi: {e}"}), 500

    return jsonify({"success": True})


@app.route("/api/export/excel")
def api_export_excel():
    user, err = require_registered()
    if err:
        return err
    db_user = get_user(user["id"])
    display_name = db_user["first_name"] or user.get("first_name", "Foydalanuvchi")

    buf = generate_excel_report(user["id"], display_name)
    filename = f"hisobchi_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/api/export/pdf")
def api_export_pdf():
    user, err = require_registered()
    if err:
        return err
    db_user = get_user(user["id"])
    display_name = db_user["first_name"] or user.get("first_name", "Foydalanuvchi")

    buf = generate_pdf_report(user["id"], display_name)
    filename = f"hisobchi_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
    return send_file(
        buf,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
    )


if __name__ == "__main__":
    init_db()
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False, threaded=True)
